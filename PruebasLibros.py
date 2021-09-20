from gutenberg.query import get_metadata
import openpyxl
import time
import wikipedia

path = "C:\\Users\\johan\\OneDrive\\Pitón\\Libros.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb['Librero']
max_row = sheet.max_row
#numlib = sheet['N1'].value
numlib = 500
cantidad = 0

while cantidad < 10:
    
    a = get_metadata('title', numlib) #Extraer datos
    b = get_metadata('author', numlib)
    c = get_metadata('subject', numlib)
    
    try:
        titulo = next(iter(a)) #Extraer Titulo
    except StopIteration:
        titulo = "Unknown Title"   
   
    try:
        autor = next(iter(b)) #Extraer Autor
    except StopIteration:
        autor = "Unknown Author"
        
    while True:               #Extraer 3 Temas
        try:
            tema = iter(c)
            tema1 = next(tema)
            tema2 = next(tema)
            tema3 = next(tema)
        except StopIteration:
            print("-")
        break
    
    sheet.cell(row=numlib, column=1).value = titulo #Escribir las celdas de Excel

    sheet.cell(row=numlib, column=2).value = autor

    sheet.cell(row=numlib, column=3).value = tema1
    
    sheet.cell(row=numlib, column=4).value = tema2
    
    sheet.cell(row=numlib, column=5).value = tema3
    
    try:
        sheet.cell(row=numlib, column=11).value = wikipedia.summary(titulo)
    except Exception as e:
        sheet.cell(row=numlib, column=11).value = "No descripción"    
       
    print (titulo)
    print (autor)
    print (tema1)
    print (tema2)
    print (tema3)
    print (numlib)

    print('\n')
    wb.save(path)
    cantidad += 1
    numlib += 1
    
    tema1 = ""
    tema2 = ""
    tema3 = ""

sheet['N1'] = numlib
wb.save(path)
print ("Terminado")