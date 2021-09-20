from bing_image_downloader import downloader
from gutenberg.query import get_metadata
from playsound import playsound
from mega import Mega
import pyshorteners
import wikipedia
import openpyxl
import time
import wget
import os


start = time.time()
#path = "E:\\OneDrive\\Piton\\Libram\\Libros.xlsx"
path = ".\\Libros.xlsx"
pathimg = "E:\\Libros"
wb = openpyxl.load_workbook(path)
sheet = wb['Librero']
max_row = sheet.max_row
numlib = sheet['N1'].value
cantidad = 0
tope = 20
mega = Mega()
fly = ""
imgfile = ""
tema1 = ""
tema2 = ""
tema3 = ""
#MegaLogin
m = mega.login("yourmail", "yourpassword")
#Shortener API
s = pyshorteners.Shortener(api_key='yourapikey', user_id='yourid',
                               domain = 0, group_id=12, type='banner')

while cantidad < tope:
    
    a = get_metadata('title', numlib) #Extraer datos
    b = get_metadata('author', numlib)
    c = get_metadata('subject', numlib)
    
    try:
        titulo = next(iter(a)) #Extraer Titulo
        bajar = 1
    except StopIteration:
        titulo = "Unknown Title"   
        bajar = 0
    try:
        autor = next(iter(b)) #Extraer Autor
    except StopIteration:
        autor = "Unknown Author"
        
    while True:               #Extraer 3 Temas
        try:
            tema = iter(c)
            tema1 = next(tema)
            tema2 = next(tema)
            tema3 = next(tema)
        except StopIteration:
            print("-")
        break

#Escribir las celdas de Excel
      
    sheet.cell(row=numlib, column=1).value = titulo 
    sheet.cell(row=numlib, column=2).value = autor
    sheet.cell(row=numlib, column=3).value = tema1    
    sheet.cell(row=numlib, column=4).value = tema2    
    sheet.cell(row=numlib, column=5).value = tema3

#Descarga imagen
    
    if bajar == 1:
        try:
            titufile = titulo.replace(':', ';')
            busqueda = str(titufile) + " by " + str(autor) + " cover book"
            downloader.download(busqueda, limit=1,  output_dir=pathimg, adult_filter_off=True, force_replace=False, timeout=60)
            os.remove(pathimg + "\\" + titufile +".jpg")
            os.rename(pathimg + "\\Image_1.jpg",pathimg + "\\" + titufile +".jpg")
            imgfile = pathimg + str(titufile) + ".jpg"
            sheet.cell(row=numlib, column=6).value = "Si img"
            sheet.cell(row=numlib, column=7).value = imgfile
        except Exception as e:
            print("No image")
            sheet.cell(row=numlib, column=6).value = "No img"
            bajar = 0
    else:
        print("No se baja libro")

#WikiDescripción
            
    if bajar == 1:
        try:
            sheet.cell(row=numlib, column=9).value = wikipedia.summary(titulo)
        except Exception as e:
            sheet.cell(row=numlib, column=9).value = "No descripción"
            bajar = 0
    else:
        print("No se baja descripción")
        bajar = 0

#Descarga Libro
    if bajar == 1:
        try:
            url = "http://aleph.gutenberg.org/cache/epub/" + str(numlib) + "/pg" + str(numlib) + "-images.epub"
            file = "E:\\Libros\\" + str(titufile) + ".epub"
            wget.download(url,file)        
            meg = m.upload(file)
            linkmega = m.get_upload_link(meg)
            fly = s.adfly.short(linkmega)
            sheet.cell(row=numlib, column=8).value = fly
        except Exception as e:
            print("No epub with images")
            
            try:
                url = "http://aleph.gutenberg.org/cache/epub/" + str(numlib) + "/pg" + str(numlib) + ".epub"
                file = "E:\\Libros\\" + str(titufile) + ".epub"
                wget.download(url,file)
                meg = m.upload(file)
                linkmega = m.get_upload_link(meg)
                fly = s.adfly.short(linkmega)
                sheet.cell(row=numlib, column=8).value = fly
            except Exception as e:
                print("No epub")
                bajar = 0
    else:
        print("No se baja libro")
        
#Mostrar Estatus
    
    print('\n')
    print ("Título: "+ titulo)
    print ("Autor: "+ autor)
    print ("Etiqueta 1: " + tema1)
    print ("Etiqueta 2: " +tema2)
    print ("Etiqueta 3: " +tema3)
    print ("Libro #"+ str(numlib))
    print ("Link AdFly: "+ fly)
    print ("Link img: "+ imgfile)
    print ("Falta(n): " + str(tope-cantidad-1) + " de " + str(tope))

    print('\n')
    wb.save(path)
    cantidad += 1
    numlib += 1
    sheet['N1'] = numlib
    
    tema1 = ""
    tema2 = ""
    tema3 = ""
    fly = ""
    imgfile = ""

sheet['N1'] = numlib
wb.save(path)
space = m.get_storage_space(giga=True)
playsound('cat.wav')
print ("Terminado")
print ("Vamos en libro: " + str(numlib))
print ("Quedan: "+ str(space) + " GB")
done = time.time()
elapsed = (done - start) / 60
print("Duración: " + str(elapsed) + " minutos.")